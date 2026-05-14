"""
reports/ フォルダの .docx・.pdf・.xlsx・.xls を読み込み、ChromaDB にインデックスを作成します。
"""
import sys
from pathlib import Path

import docx
import pypdf
import openpyxl
import chromadb
from sentence_transformers import SentenceTransformer

REPORTS_DIR = Path("reports")
CHROMA_DIR = Path("chroma_db")
COLLECTION_NAME = "dx_reports"
MODEL_NAME = "paraphrase-multilingual-mpnet-base-v2"
CHUNK_SIZE = 400   # 日本語テキストの文字数
CHUNK_OVERLAP = 80 # チャンク間オーバーラップ文字数


# ── テキスト抽出 ──────────────────────────────────────────

def read_docx(path: Path) -> str:
    doc = docx.Document(str(path))
    lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n".join(lines)


def read_pdf(path: Path) -> str:
    reader = pypdf.PdfReader(str(path))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text.strip())
    return "\n".join(pages)


def read_excel(path: Path) -> str:
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        sheet_names = wb.sheetnames
        get_rows = lambda name: wb[name].iter_rows(values_only=True)
    else:
        import xlrd
        wb = xlrd.open_workbook(str(path))
        sheet_names = [s.name for s in wb.sheets()]
        def get_rows(name):
            ws = wb.sheet_by_name(name)
            return ([ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows))

    parts = []
    for name in sheet_names:
        parts.append(f"【シート: {name}】")
        for row in get_rows(name):
            cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if cells:
                parts.append("  ".join(cells))
    return "\n".join(parts)


# ── テキスト分割 ──────────────────────────────────────────

def chunk_text(text: str) -> list[str]:
    chunks = []
    start = 0
    while start < len(text):
        end = min(start + CHUNK_SIZE, len(text))
        chunk = text[start:end].strip()
        if len(chunk) >= 15:
            chunks.append(chunk)
        if end >= len(text):
            break
        start += CHUNK_SIZE - CHUNK_OVERLAP
    return chunks


# ── メイン ────────────────────────────────────────────────

def main():
    print("=" * 50)
    print("  DX支援報告書 インデックス作成")
    print("=" * 50)

    # ── 1. ドキュメント収集（サブフォルダ含む再帰検索）──
    docx_files  = sorted(REPORTS_DIR.rglob("*.docx"))
    pdf_files   = sorted(REPORTS_DIR.rglob("*.pdf"))
    xlsx_files  = sorted(REPORTS_DIR.rglob("*.xlsx"))
    xls_files   = sorted(REPORTS_DIR.rglob("*.xls"))
    all_files   = docx_files + pdf_files + xlsx_files + xls_files

    if not all_files:
        print("\n[エラー] reports/ フォルダ（サブフォルダ含む）に .docx / .pdf / .xlsx / .xls が見つかりません")
        print("先に generate_samples.py を実行してください:")
        print("  python generate_samples.py")
        sys.exit(1)

    print(f"\n対象ファイル: {len(all_files)} 件")
    for f in all_files:
        print(f"  - {f.relative_to(REPORTS_DIR)}")

    # ── 2. テキスト抽出 & チャンク分割 ──
    print("\nテキストを抽出中...")
    all_chunks: list[str] = []
    all_metas:  list[dict] = []
    skipped_files: list[str] = []

    for file_path in all_files:
        rel_path = str(file_path.relative_to(REPORTS_DIR))
        try:
            ext = file_path.suffix.lower()
            if ext == ".docx":
                text = read_docx(file_path)
                ftype = "docx"
            elif ext == ".pdf":
                text = read_pdf(file_path)
                ftype = "pdf"
            else:
                text = read_excel(file_path)
                ftype = ext.lstrip(".")
        except Exception as e:
            print(f"  [スキップ] {rel_path}: {e}")
            skipped_files.append(rel_path)
            continue

        chunks = chunk_text(text)
        if not chunks:
            print(f"  [スキップ] {rel_path}: テキストを抽出できませんでした")
            skipped_files.append(rel_path)
            continue
        for i, chunk in enumerate(chunks):
            all_chunks.append(chunk)
            all_metas.append({
                "source":       rel_path,
                "chunk_index":  i,
                "total_chunks": len(chunks),
                "file_type":    ftype,
            })
        print(f"  {rel_path}: {len(chunks)} チャンク")

    print(f"\n合計チャンク数: {len(all_chunks)}")

    # ── 3. Embedding モデル読み込み ──
    print(f"\nEmbeddingモデルを読み込み中: {MODEL_NAME}")
    print("  (初回のみ HuggingFace から約1.1GB ダウンロードします)")
    model = SentenceTransformer(MODEL_NAME)
    print("  読み込み完了")

    # ── 4. Embedding 生成 ──
    print(f"\nEmbedding を生成中 ({len(all_chunks)} チャンク)...")
    embeddings = model.encode(
        all_chunks,
        show_progress_bar=True,
        batch_size=32,
        convert_to_numpy=True,
    )

    # ── 5. ChromaDB に登録 ──
    print("\nChromaDB にインデックスを保存中...")
    client = chromadb.PersistentClient(path=str(CHROMA_DIR))

    # 既存コレクションをリセット（再実行に対応）
    try:
        client.delete_collection(COLLECTION_NAME)
        print("  既存インデックスをリセットしました")
    except Exception:
        pass

    collection = client.create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine"},
    )

    collection.add(
        documents=all_chunks,
        embeddings=embeddings.tolist(),
        metadatas=all_metas,
        ids=[f"chunk_{i}" for i in range(len(all_chunks))],
    )

    print(f"  保存完了: {CHROMA_DIR.resolve()}")

    # ── 完了 ──
    print("\n" + "=" * 50)
    print(f"  インデックス作成完了")
    print(f"  ファイル数 : {len(all_files)} 件（スキップ: {len(skipped_files)} 件）")
    print(f"  チャンク数 : {len(all_chunks)}")
    if skipped_files:
        print("  スキップファイル:")
        for f in skipped_files:
            print(f"    - {f}")
    print("=" * 50)
    print("\n次のステップ: python search.py")
    print('  例: python search.py "製造業で在庫管理が課題の事例を探して"')


if __name__ == "__main__":
    main()
